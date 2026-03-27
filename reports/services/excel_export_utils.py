# reports/services/excel_export_utils.py

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = "366092"
TOTAL_FILL = "D3D3D3"


def create_workbook(title):
    """Создает workbook и активный лист с заданным названием."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    return wb, ws


def write_filter_row(ws, filter_info, merge_range):
    """
    Записывает строку с фильтрами.
    Возвращает стартовую строку для таблицы.
    """
    if filter_info:
        ws.merge_cells(merge_range)
        filter_cell = ws.cell(row=1, column=1, value="Фильтры: " + ", ".join(filter_info))
        filter_cell.font = Font(italic=True)
        filter_cell.alignment = Alignment(horizontal='left')
        return 3
    return 1


def write_headers(ws, row_number, headers):
    """Записывает строку заголовков таблицы."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_number, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')


def write_data_rows(ws, start_row, rows):
    """
    Записывает строки данных.
    rows: iterable со списками/кортежами значений.
    Возвращает номер следующей пустой строки.
    """
    row_idx = start_row
    for values in rows:
        for col, value in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=value)
        row_idx += 1
    return row_idx


def write_total_row(ws, row_number, values_by_col):
    """
    Записывает итоговую строку.
    values_by_col: dict {номер_колонки: значение}
    """
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color=TOTAL_FILL, end_color=TOTAL_FILL, fill_type="solid")

    for col, value in values_by_col.items():
        cell = ws.cell(row=row_number, column=col, value=value)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')


def set_column_widths(ws, widths):
    """Устанавливает ширины колонок."""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def build_excel_response(filename):
    """Создает HTTP response для xlsx-файла."""
    from django.http import HttpResponse

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response